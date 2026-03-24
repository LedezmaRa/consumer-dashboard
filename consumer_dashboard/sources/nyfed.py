"""New York Fed source adapter."""

from __future__ import annotations

from datetime import date, datetime, timezone
import hashlib
import io
import re

import httpx
from openpyxl import load_workbook

from consumer_dashboard import __version__
from consumer_dashboard.config.registry import SourceDefinition
from consumer_dashboard.sources.base import AcquisitionResult, BaseSourceAdapter
from consumer_dashboard.storage.filesystem import ensure_directory, write_json

NYFED_BASE_URL = "https://www.newyorkfed.org"
NYFED_DATABANK_URL = f"{NYFED_BASE_URL}/microeconomics/databank.html"
WORKBOOK_PATH_PATTERN = re.compile(
    r'(?P<path>/medialibrary/interactives/householdcredit/data/xls/'
    r'hhd_c_report_(?P<report_period>\d{4}q[1-4])\.xlsx(?:\?sc_lang=en)?)',
    re.IGNORECASE,
)
DATE_PATTERN = re.compile(r"(?P<release_date>[A-Za-z]+ \d{1,2}, \d{4})")
QUARTER_PATTERN = re.compile(r"^\d{2,4}:Q[1-4]$")

TARGET_SHEETS = {
    "Page 3 Data": "debt_balance_composition",
    "Page 11 Data": "balance_by_delinquency_status",
    "Page 12 Data": "ninety_plus_delinquent_balance_rate",
    "Page 13 Data": "new_delinquent_balances",
    "Page 14 Data": "new_serious_delinquent_balances",
}


def _parse_release_date(value: str) -> str:
    return datetime.strptime(value, "%B %d, %Y").date().isoformat()


def _coerce_cell_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _normalize_row(values: tuple[object, ...]) -> list[object]:
    return [_coerce_cell_value(value) for value in values]


def _extract_household_debt_release_metadata(page_text: str) -> dict[str, str] | None:
    workbook_match = WORKBOOK_PATH_PATTERN.search(page_text)
    if not workbook_match:
        return None

    workbook_path = workbook_match.group("path")
    report_period = workbook_match.group("report_period")
    context_start = max(0, workbook_match.start() - 2500)
    context = page_text[context_start:workbook_match.start()]
    date_matches = list(DATE_PATTERN.finditer(context))
    if not date_matches:
        return None

    release_date = _parse_release_date(date_matches[-1].group("release_date"))
    return {
        "release_date": release_date,
        "report_period": report_period,
        "workbook_path": workbook_path,
        "workbook_url": f"{NYFED_BASE_URL}{workbook_path}",
    }


def _extract_table_from_rows(sheet_name: str, rows: list[list[object]]) -> dict[str, object]:
    first_data_index = None
    for index, row in enumerate(rows):
        first_cell = str(row[0]).strip() if row and row[0] is not None else ""
        if QUARTER_PATTERN.match(first_cell):
            first_data_index = index
            break

    if first_data_index is None or first_data_index == 0:
        raise ValueError(f"Unable to identify quarterly data rows for sheet '{sheet_name}'.")

    header_index = first_data_index - 1
    headers = [str(value).strip() if value is not None else "" for value in rows[header_index]]

    prior_labels = []
    for row in rows[:header_index]:
        label = str(row[0]).strip() if row and row[0] is not None else ""
        if label:
            prior_labels.append(label)

    filtered_labels = [
        label
        for label in prior_labels
        if "return to table of contents" not in label.lower()
    ]
    title = filtered_labels[0] if filtered_labels else sheet_name
    unit_label = ""
    for label in filtered_labels[1:]:
        normalized = label.lower()
        if normalized.startswith("*") or "source:" in normalized:
            continue
        unit_label = label
        break
    data_rows = []

    for row in rows[first_data_index:]:
        first_cell = str(row[0]).strip() if row and row[0] is not None else ""
        if not first_cell:
            if data_rows:
                break
            continue
        if not QUARTER_PATTERN.match(first_cell):
            if data_rows:
                break
            continue

        record = {"reference_period": first_cell}
        for index, header in enumerate(headers[1:], start=1):
            if not header:
                continue
            value = row[index] if index < len(row) else None
            if value is None or value == "":
                continue
            record[header] = value
        data_rows.append(record)

    return {
        "sheet_name": sheet_name,
        "table_name": title,
        "unit_label": unit_label,
        "column_headers": [header for header in headers[1:] if header],
        "row_count": len(data_rows),
        "rows": data_rows,
    }


def _extract_selected_tables(workbook_bytes: bytes) -> dict[str, dict[str, object]]:
    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    try:
        tables = {}
        for sheet_name, table_key in TARGET_SHEETS.items():
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Workbook is missing expected sheet '{sheet_name}'.")
            sheet = workbook[sheet_name]
            rows = [_normalize_row(row) for row in sheet.iter_rows(values_only=True)]
            tables[table_key] = _extract_table_from_rows(sheet_name, rows)
        return tables
    finally:
        workbook.close()


class NewYorkFedSourceAdapter(BaseSourceAdapter):
    source_id = "new_york_fed"

    def fetch_latest(self, definition: SourceDefinition) -> AcquisitionResult:
        fetched_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        artifact_dir = ensure_directory(self.settings.raw_dir / self.source_id / run_id)
        artifact_path = artifact_dir / "household_debt_credit_report.json"

        try:
            with httpx.Client(timeout=self.settings.http_timeout_seconds) as client:
                page_response = client.get(NYFED_DATABANK_URL)
                page_response.raise_for_status()
                page_text = page_response.text

                metadata = _extract_household_debt_release_metadata(page_text)
                if metadata is None:
                    return AcquisitionResult(
                        source_id=self.source_id,
                        status="parse_failed",
                        message=(
                            "Unable to identify the latest Household Debt and Credit workbook "
                            "from the New York Fed databank page."
                        ),
                    )

                workbook_response = client.get(metadata["workbook_url"])
                workbook_response.raise_for_status()
                workbook_bytes = workbook_response.content
        except httpx.HTTPError as exc:
            return AcquisitionResult(
                source_id=self.source_id,
                status="request_failed",
                message=f"New York Fed request failed: {exc}",
            )

        try:
            tables = _extract_selected_tables(workbook_bytes)
        except Exception as exc:
            return AcquisitionResult(
                source_id=self.source_id,
                status="parse_failed",
                message=f"Unable to parse the New York Fed Household Debt and Credit workbook: {exc}",
            )

        envelope = {
            "metadata": {
                "source_id": self.source_id,
                "source_name": definition.source_name,
                "report_slug": "household_debt_credit",
                "report_name": "Household Debt and Credit",
                "fetched_at": fetched_at,
                "release_date": metadata["release_date"],
                "report_period": metadata["report_period"],
                "artifact_type": "xlsx_extract",
                "adapter_version": __version__,
                "endpoint": metadata["workbook_url"],
                "catalog_url": NYFED_DATABANK_URL,
                "http_status": workbook_response.status_code,
                "page_http_status": page_response.status_code,
                "response_sha256": hashlib.sha256(workbook_bytes).hexdigest(),
                "selected_sheets": list(TARGET_SHEETS),
            },
            "tables": tables,
        }
        write_json(artifact_path, envelope)

        total_rows = sum(table.get("row_count", 0) for table in tables.values())
        return AcquisitionResult(
            source_id=self.source_id,
            status="ingested",
            message=(
                "Downloaded New York Fed Household Debt and Credit data for "
                f"{metadata['report_period'].upper()} released on {metadata['release_date']} "
                f"({total_rows} quarterly rows across {len(tables)} extracted tables)."
            ),
            artifacts=[str(artifact_path)],
        )
